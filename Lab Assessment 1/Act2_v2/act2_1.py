import paho.mqtt.client as mqtt
import copy
import time
import openpyxl
import json

# Callback when the client connects to the MQTT broker
def on_connect(client, userdata, flags, rc):
    if rc == 0:
        print("Connected to MQTT broker\n")
    else:
        print("Connection failed with code {rc}")


# Create an MQTT client instance
client = mqtt.Client("")

# Set the callback function
client.on_connect = on_connect

broker_address = "test.mosquitto.org"  # broker's address
broker_port = 1883
keepalive = 5
qos = 2
publish_topic = "Protocol_pros_1"

# Connect to the MQTT broker
client.connect(broker_address, broker_port, keepalive)

# Start the MQTT loop to handle network traffic
client.loop_start()

# Publish loop

try:
    while True:
        # Publish a message to the send topic
        #Accessing the excel datasheet 
        dataframe = openpyxl.load_workbook("Location data.xlsx")
        dataframe1 = dataframe.active

        #Getting the topic row
        topics = []
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            topics.append(col[0].value) 
        #print(topics)

        #To index the locations
        i = 0
        #To collect the data
        data = {}
        #To temporary collect data
        temp = {}
        for row in dataframe1.iter_rows(2, dataframe1.max_row):
            for col in range(0, dataframe1.max_column):
                temp[topics[col]] = str(row[col].value)
            insert = copy.deepcopy(temp)
            data[str(i)] = insert 
            temp.clear()   
            i += 1
        print(data)

        # Serializing json  
        json_object = json.dumps(data, indent = 4) 
        
        value = json_object
        client.publish(publish_topic,value,qos)
        print(f"Published message '{value}' to topic '{publish_topic}'\n")
        
        # Wait for a moment to simulate some client activity
        time.sleep(5)

except KeyboardInterrupt:
    # Disconnect from the MQTT broker
    pass
client.loop_stop()
client.disconnect()

print("Disconnected from the MQTT broker")
